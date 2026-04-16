"""
Run once to seed Supabase with historical data from Expenditure.xlsx.
Usage:
    cd backend
    python seed.py
"""

import asyncio
import os
import datetime
from pathlib import Path
import openpyxl
from dotenv import load_dotenv

load_dotenv()

# must be set before importing app modules
from database import get_session_factory, create_tables
from models import Group, Member, Expense

XLSX_PATH = Path(__file__).parent.parent / "Expenditure.xlsx"

# ── Sheet → trip metadata ────────────────────────────────────────────────────
SHEET_META = {
    "Sheet1": {
        "name": "Friends Outings",
        "emoji": "🍻",
        "description": "Ajay, Anukul & Anubhav – Mar to Jun 2025",
        "members": ["Ajay", "Anukul", "Anubhav"],
    },
    "Sheet4": {
        "name": "Anubhav & Me",
        "emoji": "🤝",
        "description": "Anubhav & Anukul – Mar to Jun 2025",
        "members": ["Anubhav", "Anukul"],
    },
    "Sheet5": {
        "name": "Goa / Mumbai Trip",
        "emoji": "🌊",
        "description": "Anukul, Apoorv & Akshat – Oct 3–5, 2025",
        "members": ["Anukul", "Apoorv", "Akshat"],
    },
    "Sheet6": {
        "name": "Mumbai with Apoorv",
        "emoji": "🎬",
        "description": "Apoorv & Anukul – Oct 1–2, 2025",
        "members": ["Apoorv", "Anukul"],
    },
    "Sheet8": {
        "name": "Mumbai + Diwali Trip",
        "emoji": "🪔",
        "description": "Anubhav & Anukul – Nov 2025",
        "members": ["Anubhav", "Anukul"],
    },
    "Sheet2": {
        "name": "Jupyter Card (Personal)",
        "emoji": "💳",
        "description": "Personal card expenses – Apr–May 2025",
        "members": ["Anukul"],
    },
}


def _parse_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    raw = str(value).strip()
    # try DD-MM-YY
    for fmt in ("%d-%m-%y", "%d-%m-%Y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw


def _parse_amount(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _load_expenses(ws, members: list[str]) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    # column index helpers
    def col(name: str) -> int | None:
        for i, h in enumerate(header):
            if name in h:
                return i
        return None

    ci_date = col("date")
    ci_type = col("type")
    ci_title = col("title")
    ci_amount = col("amount")
    ci_name = col("name")
    ci_div = col("divider")
    ci_ind = col("individual")

    expenses = []
    for row in rows[1:]:
        # stop at summary / blank rows
        if all(v is None for v in row):
            continue
        amount_raw = row[ci_amount] if ci_amount is not None else None
        amount = _parse_amount(amount_raw)
        if amount is None or amount <= 0:
            continue

        paid_by = str(row[ci_name]).strip() if ci_name is not None and row[ci_name] else ""
        if not paid_by or paid_by.lower() in ("none", "paid", "p.p", "indivdual", "individual"):
            continue

        divider_raw = row[ci_div] if ci_div is not None else None
        divider = int(_parse_amount(divider_raw) or len(members))
        individual_raw = row[ci_ind] if ci_ind is not None else None
        individual = _parse_amount(individual_raw) if individual_raw is not None else round(amount / divider, 2)

        expenses.append({
            "date": _parse_date(row[ci_date] if ci_date is not None else None),
            "category": str(row[ci_type]).strip() if ci_type is not None and row[ci_type] else None,
            "title": str(row[ci_title]).strip() if ci_title is not None and row[ci_title] else None,
            "amount": round(amount, 2),
            "paid_by": paid_by,
            "divider": divider,
            "individual_amount": round(individual, 2) if individual else None,
        })

    return expenses


def _load_sheet2(ws) -> list[dict]:
    """Sheet2 is a simple list of amounts – no payer column, treated as personal."""
    expenses = []
    rows = list(ws.iter_rows(values_only=True))
    date_label = None
    for row in rows[1:]:
        v = row[0]
        if v is None:
            continue
        if isinstance(v, str):
            date_label = v   # e.g. "16apr-16may"
            continue
        amount = _parse_amount(v)
        if amount and amount > 0:
            expenses.append({
                "date": None,
                "category": date_label,
                "title": None,
                "amount": round(amount, 2),
                "paid_by": "Anukul",
                "divider": 1,
                "individual_amount": round(amount, 2),
            })
    return expenses


async def seed():
    await create_tables()
    factory = get_session_factory()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    async with factory() as db:
        for sheet_name, meta in SHEET_META.items():
            if sheet_name not in wb.sheetnames:
                print(f"  skip {sheet_name} (not found)")
                continue

            print(f"\nSeeding {sheet_name} -> '{meta['name']}'")

            group = Group(
                name=meta["name"],
                description=meta["description"],
                emoji=meta["emoji"],
                is_historical=True,
            )
            db.add(group)
            await db.flush()

            for name in meta["members"]:
                db.add(Member(group_id=group.id, name=name))

            ws = wb[sheet_name]
            expense_rows = (
                _load_sheet2(ws) if sheet_name == "Sheet2"
                else _load_expenses(ws, meta["members"])
            )

            for er in expense_rows:
                db.add(Expense(
                    group_id=group.id,
                    date=er["date"],
                    category=er["category"],
                    title=er["title"],
                    amount=er["amount"],
                    paid_by=er["paid_by"],
                    divider=er["divider"],
                    individual_amount=er["individual_amount"],
                ))

            print(f"  {len(expense_rows)} expenses added")

        await db.commit()
        print("\nDone.")


if __name__ == "__main__":
    asyncio.run(seed())
